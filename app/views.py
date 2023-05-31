from django.shortcuts import render
from .models import Items
from rest_framework.response import Response
from rest_framework.decorators import api_view,permission_classes
from .serializer import ProductSerializers
from .models import Items
from rest_framework import status
from django.http import HttpResponse
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from datetime import date
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from rest_framework.permissions import IsAuthenticated
from openpyxl import Workbook
from openpyxl.styles import PatternFill,Border, Side,Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
import matplotlib.pyplot as plt
import io
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.barcharts import VerticalBarChart
from matplotlib.backends.backend_pdf import PdfPages

from reportlab.graphics.charts.legends import Legend



@api_view(["GET"])
def api_Overview(req):
    over_view ={
        'list_all':'all/'
    }
    return Response({"message":"Api Works","data":over_view})


@api_view(["GET"])
def get_all(req):
    products = Items.objects.all()
    serializer = ProductSerializers(products,many=True)
    print(serializer)
    return Response(serializer.data)

@api_view(["POST"])
def createItem(req):
    user = ProductSerializers(data=req.data)
    if user.is_valid():
        user.save()
    return Response({"message":"User Create SuccessFully"},status=status.HTTP_201_CREATED)

@api_view(['POST'])
def updateItem(req, pk):
    item = Items.objects.get(id=pk)
    itemSerializer = ProductSerializers(instance=item, data=req.data)
    if itemSerializer.is_valid():
        itemSerializer.save()
        return Response(itemSerializer.data)
    

@api_view(["GET",])
def pdf_gen(req):
    items = Items.objects.all()
    response = HttpResponse(content_type="application/pdf")
    response['Content-Disposition'] = 'attachment; filename="generated_pdf.pdf"'
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    heading_style = ParagraphStyle(
        name='Products',
        parent=styles['Normal'],
        fontSize=14,
        textColor=colors.black,
        spaceAfter=12,
        alignment=1,
    )

    heading_text = 'Products'
    heading_paragraph = Paragraph(heading_text, heading_style)
    elements.append(heading_paragraph)

    # Create chart data
    chart_data = [
        [1000, 500, 700, 1000]
    ]
    drawing = Drawing(600, 300)
    chart = VerticalBarChart()
    chart.x = 20
    chart.y = 0
    chart.width = 400
    chart.height = 200
    chart.data = chart_data
    chart.strokeColor = colors.black

    # Add the chart to the drawing
    legend = Legend()
    legend.x = 350
    legend.y = 10
    legend.boxAnchor = 'ne'  # Set the position of the legend box
    legend.columnMaximum = 1  # Set the maximum number of columns in the legend
    legend.colorNamePairs = [(colors.red, 'Name'),(colors.blue, 'Se'), (colors.blue, 'Sem'),(colors.green,"Value")]
    legend.fontName = 'Helvetica'
    legend.fontSize = 10

    drawing.add(legend)
    drawing.add(chart)

    # Add the drawing to the elements list
    elements.append(drawing)

    # Create table data
    table_data = [['Name', 'Description']]
    for item in items:
        row = [str(item.name), str(item.description)]
        table_data.append(row)

    # Define table styles
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.aliceblue),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.aliceblue),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])

    # Create the table
    table = Table(table_data, repeatRows=1, colWidths=[200, 200])
    table.setStyle(table_style)
    doc.build(elements)
    return response


@api_view(['GET'])
def gen_excel(request):
    thin = Side(border_style="thin", color="00000000")
    border_style =  Border(top=thin, left=thin, right=thin, bottom=thin)
    items = Items.objects.all()
    excel = Workbook()
    sheet = excel.active
    cell1 = sheet['A1']
    cell2 = sheet["B1"]
    cell1.fill = PatternFill(start_color="00FFFF00", end_color="00FFFF00", fill_type="solid")
    cell2.fill = PatternFill(start_color="00FFFF00", end_color="00FFFF00", fill_type="solid")
    cell1.border = border_style
    cell2.border = border_style
    cell1.alignment = Alignment(horizontal="center", vertical="center")
    cell2.alignment = Alignment(horizontal="center", vertical="center")

    sheet['A1'] = "Name"
    sheet["B1"] = "Description"

    for index, item in enumerate(items, start=2):
        sheet[f'A{index}'].border = border_style
        sheet[f'B{index}'].border = border_style
        sheet[f'A{index}'] = item.name
        sheet[f'B{index}'] = item.description

    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2.2)
        sheet.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename=dynamic_excel.xlsx'
    excel.save(response)
    return response